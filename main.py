import openpyxl as op
import re
from func import *

# load database with about 1 200 examples
filename = "data.xlsx"
db = op.load_workbook(filename)
ws = db["Sheet1"]		 #contains only right examples, used for IE_finder()
ws_wrong = db["Sheet2"]  #contains only examples with mistakes

# can modify this sentence and search for mistakes in it, right now there is missing '(' before 'véase'
trial_sentence = "human being(s)_g véase también human rights; Universal Declaration)"


def main(sentence):

    phrase = re.sub("\d", "*", sentence).lower()

    if "véase" in phrase:
        CPR = phrase[phrase.find("véase")-1:]
        IE_NG = phrase[:phrase.find("véase")-1]

        print(trouble_killer(phrase, IE_NG, CPR, ws))


    elif "véanse" in phrase:
        CPR = phrase[phrase.find("véanse")-1:]
        IE_NG = phrase[:phrase.find("véanse")-1]

        print(trouble_killer(phrase, IE_NG, CPR, ws))


    else:
        syntax_errors.append(errors["véase_not_err"])
        print(syntax_errors)



# code to test just trial_sentence
main(trial_sentence) 




# loop through all the examples in Sheet1 of data.xlsx
# UNCOMMENT TO LAUNCH!
# for i in range(1, ws.max_row + 1):

#     if ws.cell(column = 1, row = i).value != None:
#         sentence = ws.cell(column = 1, row = i).value

#         main(sentence)
#     else:
#         print("passed")



# loop through all the WRONG examples in Sheet2 of data.xlsx, made to find mistakes
# I tried to get results to the xlsx sheet but was unable to do that so you need to read mistakes from terminal
# UNCOMMENT TO LAUNCH!  
# for i in range(1, ws_wrong.max_row + 1):

#     if ws_wrong.cell(column = 1, row = i).value != None:
#         sentence = ws_wrong.cell(column = 1, row = i).value
   
#         print(i)          #shows Sheet index of the sentence so you can find it
#         main(sentence)    #shows mistakes in terminal
        
#     else:
#         print("passed")


